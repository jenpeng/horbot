"""
Excel MCP Server - Excel 文件操作工具

提供 Excel 文件的创建、读取、写入等功能。
"""

import os
from typing import Any, Optional
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.types import Tool, TextContent

# 尝试导入 openpyxl
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# 创建 MCP Server 实例
server = Server("excel-server")


def check_openpyxl() -> str | None:
    """检查 openpyxl 是否可用"""
    if not HAS_OPENPYXL:
        return "❌ openpyxl 未安装。请运行: pip install openpyxl"
    return None


@server.list_tools()
async def list_tools() -> list[Tool]:
    """列出所有可用工具"""
    return [
        Tool(
            name="excel_create_workbook",
            description="创建新的 Excel 工作簿",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "文件保存路径（如 /path/to/file.xlsx）"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称（可选，默认为 Sheet）",
                        "default": "Sheet"
                    }
                },
                "required": ["filepath"]
            }
        ),
        Tool(
            name="excel_write_data",
            description="向 Excel 文件写入数据（支持单格和区域）",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Excel 文件路径"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称（可选，默认为活动工作表）"
                    },
                    "data": {
                        "type": "array",
                        "description": "数据数组，每行是一个数组。如 [['A1', 'B1'], ['A2', 'B2']]",
                        "items": {
                            "type": "array",
                            "items": {
                                "type": ["string", "number", "boolean", "null"]
                            }
                        }
                    },
                    "start_cell": {
                        "type": "string",
                        "description": "起始单元格（如 A1），默认为 A1",
                        "default": "A1"
                    }
                },
                "required": ["filepath", "data"]
            }
        ),
        Tool(
            name="excel_read_data",
            description="读取 Excel 文件中的数据",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Excel 文件路径"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称（可选，默认为活动工作表）"
                    },
                    "range": {
                        "type": "string",
                        "description": "读取范围（如 A1:C10），不指定则读取所有数据"
                    }
                },
                "required": ["filepath"]
            }
        ),
        Tool(
            name="excel_set_style",
            description="设置单元格样式（字体、背景色、边框等）",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Excel 文件路径"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "工作表名称"
                    },
                    "range": {
                        "type": "string",
                        "description": "样式应用范围（如 A1:E1）"
                    },
                    "bold": {
                        "type": "boolean",
                        "description": "是否加粗"
                    },
                    "font_size": {
                        "type": "integer",
                        "description": "字体大小"
                    },
                    "bg_color": {
                        "type": "string",
                        "description": "背景颜色（十六进制，如 FFFF00）"
                    },
                    "align": {
                        "type": "string",
                        "description": "对齐方式：left/center/right",
                        "enum": ["left", "center", "right"]
                    }
                },
                "required": ["filepath", "range"]
            }
        ),
        Tool(
            name="excel_list_sheets",
            description="列出 Excel 文件中的所有工作表",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Excel 文件路径"
                    }
                },
                "required": ["filepath"]
            }
        ),
        Tool(
            name="excel_add_sheet",
            description="添加新工作表",
            inputSchema={
                "type": "object",
                "properties": {
                    "filepath": {
                        "type": "string",
                        "description": "Excel 文件路径"
                    },
                    "sheet_name": {
                        "type": "string",
                        "description": "新工作表名称"
                    }
                },
                "required": ["filepath", "sheet_name"]
            }
        )
    ]


@server.call_tool()
async def call_tool(name: str, arguments: dict[str, Any]) -> list[TextContent]:
    """执行工具调用"""
    
    # 检查依赖
    error = check_openpyxl()
    if error:
        return [TextContent(type="text", text=error)]
    
    try:
        if name == "excel_create_workbook":
            return await create_workbook(arguments)
        elif name == "excel_write_data":
            return await write_data(arguments)
        elif name == "excel_read_data":
            return await read_data(arguments)
        elif name == "excel_set_style":
            return await set_style(arguments)
        elif name == "excel_list_sheets":
            return await list_sheets(arguments)
        elif name == "excel_add_sheet":
            return await add_sheet(arguments)
        else:
            return [TextContent(type="text", text=f"❌ 未知工具: {name}")]
    except Exception as e:
        return [TextContent(type="text", text=f"❌ 执行失败: {str(e)}")]


async def create_workbook(args: dict) -> list[TextContent]:
    """创建新工作簿"""
    filepath = args["filepath"]
    sheet_name = args.get("sheet_name", "Sheet")
    
    # 确保目录存在
    dir_path = os.path.dirname(filepath)
    if dir_path:
        os.makedirs(dir_path, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    wb.save(filepath)
    
    return [TextContent(
        type="text",
        text=f"✅ 工作簿创建成功！\n📁 路径: {filepath}\n📋 工作表: {sheet_name}"
    )]


async def write_data(args: dict) -> list[TextContent]:
    """写入数据"""
    filepath = args["filepath"]
    data = args["data"]
    sheet_name = args.get("sheet_name")
    start_cell = args.get("start_cell", "A1")
    
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    # 计算起始行列
    start_col = ord(start_cell[0].upper()) - ord('A') + 1
    start_row = int(''.join(filter(str.isdigit, start_cell)))
    
    # 写入数据
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)
    
    wb.save(filepath)
    
    return [TextContent(
        type="text",
        text=f"✅ 数据写入成功！\n📁 文件: {filepath}\n📊 范围: {start_cell} 起，共 {len(data)} 行 {len(data[0]) if data else 0} 列"
    )]


async def read_data(args: dict) -> list[TextContent]:
    """读取数据"""
    filepath = args["filepath"]
    sheet_name = args.get("sheet_name")
    cell_range = args.get("range")
    
    wb = load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    if cell_range:
        # 读取指定范围
        data = []
        for row in ws[cell_range]:
            row_data = [cell.value for cell in row]
            data.append(row_data)
    else:
        # 读取所有数据
        data = []
        for row in ws.iter_rows():
            row_data = [cell.value for cell in row]
            if any(v is not None for v in row_data):  # 跳过空行
                data.append(row_data)
    
    # 格式化输出
    result = f"📖 读取数据（工作表: {ws.title}）:\n\n"
    for row in data:
        result += " | ".join(str(v) if v is not None else "" for v in row) + "\n"
    
    return [TextContent(type="text", text=result)]


async def set_style(args: dict) -> list[TextContent]:
    """设置样式"""
    filepath = args["filepath"]
    cell_range = args["range"]
    sheet_name = args.get("sheet_name")
    
    wb = load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active
    
    # 构建样式
    font_kwargs = {}
    if args.get("bold"):
        font_kwargs["bold"] = True
    if args.get("font_size"):
        font_kwargs["size"] = args["font_size"]
    
    fill = None
    if args.get("bg_color"):
        fill = PatternFill(start_color=args["bg_color"], end_color=args["bg_color"], fill_type="solid")
    
    alignment = None
    if args.get("align"):
        alignment = Alignment(horizontal=args["align"])
    
    # 应用样式
    for row in ws[cell_range]:
        for cell in row:
            if font_kwargs:
                cell.font = Font(**font_kwargs)
            if fill:
                cell.fill = fill
            if alignment:
                cell.alignment = alignment
    
    wb.save(filepath)
    
    return [TextContent(
        type="text",
        text=f"✅ 样式设置成功！\n📁 文件: {filepath}\n📊 范围: {cell_range}"
    )]


async def list_sheets(args: dict) -> list[TextContent]:
    """列出所有工作表"""
    filepath = args["filepath"]
    
    wb = load_workbook(filepath)
    sheets = wb.sheetnames
    
    result = f"📋 工作表列表:\n"
    for i, name in enumerate(sheets, 1):
        result += f"  {i}. {name}\n"
    
    return [TextContent(type="text", text=result)]


async def add_sheet(args: dict) -> list[TextContent]:
    """添加工作表"""
    filepath = args["filepath"]
    sheet_name = args["sheet_name"]
    
    wb = load_workbook(filepath)
    wb.create_sheet(title=sheet_name)
    wb.save(filepath)
    
    return [TextContent(
        type="text",
        text=f"✅ 工作表添加成功！\n📁 文件: {filepath}\n📋 新工作表: {sheet_name}"
    )]


async def main():
    """启动 MCP Server"""
    async with stdio_server() as (read_stream, write_stream):
        await server.run(
            read_stream,
            write_stream,
            server.create_initialization_options()
        )


if __name__ == "__main__":
    import asyncio
    asyncio.run(main())